import os
import sys
import openpyxl
import datetime as dt
from openpyxl.styles import Font, Border, Side, NamedStyle, PatternFill, Color, colors, Alignment
from openpyxl.utils import get_column_letter
import win32com.client as win32

def kitchen(kitchen_filename):
    print('''
     _  __  _   _          _                 
    | |/ / (_) | |_   __  | |_    ___   _ _  
    | ' <  | | |  _| / _| | ' \\  / -_) | ' \\ 
    |_|\\_\\ |_|  \\__| \\__| |_||_| \\___| |_||_|
    ''')

    # Открытие файла
    file_name = kitchen_filename + 'x'
    file_path = os.getcwd() + '\\' + kitchen_filename

    # Конвертирование расширение .xls -> .xlsx
    fname = file_path
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)
    wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
    wb.Close()                               #FileFormat = 56 is for .xls extension
    excel.Application.Quit()

    print(f'[Кухня] Открываю файл {file_name}.')
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.title = 'Лист кухни'

    # Назначение даты файла
    day_yesterday = (dt.date(int(ws['B2'].value[6:10]), int(ws['B2'].value[3:5]), int(ws['B2'].value[:2])) - dt.timedelta(1)).strftime('%d.%m.%Y')
    file_date = ws['B2'].value

    # Редактирование
    ws.unmerge_cells('A4:B4')
    ws.delete_rows(1)
    ws['B3'] = ws['A3'].value   # Перенос 'Номер'
    ws['B' + str(ws.max_row)] = ''
    for row_num in range(3, ws.max_row):      # Очистка столбца 'A#'
        ws['A' + str(row_num)] = ''

    # Подсчёт гостей
    for row_num in range(4, ws.max_row+1):
        try:      
            ws['C' + str(row_num)] = int(ws['D' + str(row_num)].value) + int(ws['E' + str(row_num)].value)
        except:
            ws['C' + str(row_num)] = ''
        try:
            ws['F' + str(row_num)] = int(ws['G' + str(row_num)].value) + int(ws['H' + str(row_num)].value)
        except:
            ws['F' + str(row_num)] = ''
        try:
            ws['I' + str(row_num)] = int(ws['J' + str(row_num)].value) + int(ws['K' + str(row_num)].value)
        except:
            ws['I' + str(row_num)] = ''

    # Удаление лишних ячеек
    useless_columns = [4, 5, 7, 8, 10, 11]
    for column_num in reversed(range(1, ws.max_column + 1)):
        if column_num in useless_columns:
            ws.delete_cols(column_num)

    for row_num in range(4, ws.max_row):
        if ws['C' + str(row_num)].value == '' and ws['D' + str(row_num)].value == '' and ws['E' + str(row_num)].value == '':
            ws.delete_rows(row_num)

    # Пересохранение файла потому что вылезает ошибка, что строк больше, чем нужно
    # (баг возникает при наличии позднего/раннего выезда/заезда)
    print('[Кухня] Пересохранение.')
    file_name = f'Kitchen_{file_date}.xlsx'
    wb.save(file_name)
    print(f'[Кухня] Открытие {file_name}.')
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active


    # Подсчёт СЕГОДНЯ + ЗАВТРА 1
    today = ws['E' + str(ws.max_row)].value + ws['C' + str(ws.max_row)].value
    tomorrow = ws['E' + str(ws.max_row)].value + ws['D' + str(ws.max_row)].value
    guest_today = 'ГОСТ'
    guest_tomorrow = 'ГОСТ'

    if today in range(5, 21):
        guest_today += 'ЕЙ'
    elif today % 10 == 1:
        guest_today += 'Ь'
    elif 2 <= today % 10 <= 4:
        guest_today += 'Я'
    else:
        guest_today += 'ЕЙ'

    if tomorrow in range(5, 21):
        guest_tomorrow += 'ЕЙ'
    elif tomorrow % 10 == 1:
        guest_tomorrow += 'Ь'
    elif 2 <= tomorrow % 10 <= 4:
        guest_tomorrow += 'Я'
    else:
        guest_tomorrow += 'ЕЙ'

    # Редактирование
    ws['E3'] = 'Прожив'
    ws['A' + str(ws.max_row)] = 'Номеров:' + str(ws.max_row - 4)
    ws['C' + str(ws.max_row)] = 'Выезд:' + str(ws['C' + str(ws.max_row)].value)
    ws['D' + str(ws.max_row)] = 'Заезд:' + str(ws['D' + str(ws.max_row)].value)
    ws['E' + str(ws.max_row)] = 'Прожив:' + str(ws['E' + str(ws.max_row)].value)

    # Подсчёт СЕГОДНЯ + ЗАВТРА 2
    ws['A' + str(ws.max_row+2)] = f'СЕГОДНЯ: {today} {guest_today}'
    ws['A' + str(ws.max_row+1)] = f'ЗАВТРА: ~{tomorrow} {guest_tomorrow}'

    # Ширина столбцов
    column_letter = ['B', 'C', 'D', 'E']
    column_width = [7.14, 9.28, 8.71, 10.71]
    for i in range(len(column_letter)):
        ws.column_dimensions[column_letter[i]].width = column_width[i]
    # Ширина строк
    for i in range(3, ws.max_row-3):
        ws.row_dimensions[i].height = 11.25

    # Обрамление рамкой
    dotted = Side(border_style='dotted', color='000000')
    for column_num in range(2, ws.max_column+1):
            for row_num in range(3, ws.max_row-3):
                ws.cell(row=row_num, column=column_num).border = Border(top=dotted, left=dotted, right=dotted, bottom=dotted)
                ws.cell(row=row_num, column=column_num).alignment = Alignment(horizontal="center", vertical="center")
    # Размер шрифта
    ws['A1'].font = Font(size = 20)
    ws['B1'].font = Font(size = 20)
    ws['A' + str(ws.max_row)].font = Font(size = 20)
    ws['A' + str(ws.max_row-1)].font = Font(size = 20)

    # ________Если есть файл за предыдущий день________
    if f'Kitchen_{day_yesterday}.xlsx' in os.listdir():
        print('[Кухня] Найден отчёт за вчерашний день. Переношу старый отчёт в папку old.')
        file_name_last = f'Kitchen_{day_yesterday}.xlsx'
        file_path_last = os.getcwd() + '\\' + file_name_last
        try:
            os.rename(file_path_last, os.getcwd() + '\\old\\' + file_name_last)
        except:
            os.remove(file_path_last)

    # Сохранение файла
    print('[Кухня] Сохраняю файл.')
    wb.save(f'Kitchen_{file_date}.xlsx')

    print(f'[Кухня] ГОТОВО. Лист кухни сохранён под именем "Kitchen_{file_date}".')

    os.remove(file_path)
    os.remove(file_path + 'x')

def checkin(checkin_filename):
    print('''
     ___   _                 _           _        
    / __| | |_    ___   __  | |__  ___  (_)  _ _  
   | (__  | ' \\  / -_) / _| | / / |___| | | | ' \\ 
    \\___| |_||_| \\___| \\__| |_\\_\\       |_| |_||_|
    ''')

    # Открытие файла
    file_name = checkin_filename
    file_path = os.getcwd() + '\\' + checkin_filename

    print(f'[Заезд] Открываю файл {file_name}.')
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.title = 'Лист заезда'

    # Назначение даты файла
    day_yesterday = (dt.date(int(ws['E2'].value[6:10]), int(ws['E2'].value[3:5]), int(ws['E2'].value[:2])) - dt.timedelta(1)).strftime('%d.%m.%Y')
    file_date = ws['E2'].value[:10]

    # Удаление лишних ячеек
    useless_columns = [1, 3, 4, 7, 8, 11, 13]

    for column_num in reversed(range(1, ws.max_column + 1)):
        if column_num in useless_columns:
            ws.delete_cols(column_num)
    ws['H1'] = 'Примечания'

    # Ширина столбцов
    column_letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    column_width = [9.14, 5, 5, 8.57, 22.86, 4.29, 6.29, 29.29]
    for i in range(len(column_letter)):
        ws.column_dimensions[column_letter[i]].width = column_width[i]

    # Обрамление рамкой
    thin = Side(border_style='thin', color='000000')
    thick = Side(border_style='thick', color='000000')

    for column_num in range(1, ws.max_column + 1):
            for row_num in range(1, ws.max_row + 1):
                ws.cell(row=row_num, column=column_num).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 5 дополнительных ячеек для новых броней
    additional_row = ws.max_row
    for column_num in range(1, ws.max_column + 1):
            for row_num in range(additional_row + 1, additional_row + 6):
                ws.cell(row=row_num, column=column_num).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for i in range(ws.max_row-4, ws.max_row+1):
        ws.row_dimensions[i].height = 18.75

    # Жирный шрифт верхней строчке + жирная полоска под
    for column_num in range(1, ws.max_column + 1):
        ws[get_column_letter(column_num) + '1'].font = Font(bold = True)
        ws[get_column_letter(column_num) + '1'].border = Border(top=thin, left=thin, right=thin, bottom=thick)

    # Сортировка по номеру
    ws.auto_filter.ref = "D1:D100"
    ws.auto_filter.add_sort_condition("D2:D100")

    # Серый цвет через строчку
    greyFill = PatternFill(start_color='FFDDDDDD',
                        end_color='FFDDDDDD',
                        fill_type='solid')

    rooms = []
    for row_num in range(2, ws.max_row - 4):
        rooms.append(ws['D' + str(row_num)].value)
    rooms.sort()

    for i in range(len(rooms)):
        if i % 2 == 1:
            for row_num in range(2, ws.max_row + 1):
                if ws['D' + str(row_num)].value == rooms[i]:
                    for column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                        ws[column_letter + str(row_num)].fill = greyFill

    # Если есть РАННИЙ ЗАЕЗД
    for row_num in range(2, ws.max_row - 4):
        if '14:00' not in ws['B' + str(row_num)].value:
            ws['I' + str(row_num)].value = 'РЗ'

    # ________Если есть файл за предыдущий день________
    if f'Check-in_{day_yesterday}.xlsx' in os.listdir():
        print('[Заезд] Найден отчёт за вчерашний день. Переношу старый отчёт в папку old.')
        file_name_last = f'Check-in_{day_yesterday}.xlsx'
        file_path_last = os.getcwd() + '\\' + file_name_last
        try:
            os.rename(file_path_last, os.getcwd() + '\\old\\' + file_name_last)
        except:
            os.remove(file_path_last)

    # Сохранение файла
    print('[Заезд] Сохраняю файл.')
    wb.save(f'Check-in_{file_date}.xlsx')

    os.remove(file_path)

    print(f'[Заезд] ГОТОВО. Лист заезда сохранён под именем "Check-in_{file_date}".')

def living(living_filename):
    print('''
     _      _         _               
    | |    (_) __ __ (_)  _ _    __ _ 
    | |__  | | \\ V / | | | ' \\  / _` |
    |____| |_|  \\_/  |_| |_||_| \\__, |
                                |___/ 
    ''')

    # Открытие файла
    file_name = living_filename
    file_path = os.getcwd() + '\\' + living_filename

    print(f'[Проживание] Открываю файл {file_name}.')
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.title = 'Лист проживающих'

    # Назначение даты файла
    day_today = dt.date.today().strftime('%d.%m.%Y')
    day_tomorrow = (dt.date.today() + dt.timedelta(1)).strftime('%d.%m.%Y')
    day_yesterday = (dt.date.today() - dt.timedelta(1)).strftime('%d.%m.%Y')
    
    print(f'\nНа какое число делаем проживание? \n\n\
        Введите [1] если на сегодня {day_today}. \n\
        Введите [2] если на завтра {day_tomorrow}.')

    n = int(input('\nВведите цифру и нажмите ENTER: '))

    if n == 1:
        file_date = day_today
        file_date_last = day_yesterday
    elif n == 2:
        file_date = day_tomorrow
        file_date_last = day_today

    # Удаление лишних ячеек
    useless_columns = [1, 3, 4, 7, 8, 11, 13]

    for column_num in reversed(range(1, ws.max_column + 1)):
        if column_num in useless_columns:
            ws.delete_cols(column_num)
    ws['H1'] = 'мб'
    ws['I1'] = 'Примечания'

    # Ширина столбцов
    column_letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    column_width = [9.14, 5, 5, 7.14, 21.43, 4.29, 6.29, 6, 26.14]
    for i in range(len(column_letter)):
        ws.column_dimensions[column_letter[i]].width = column_width[i]

    # Обрамление рамкой
    thin = Side(border_style='thin', color='000000')
    thick = Side(border_style='thick', color='000000')

    for column_num in range(1, ws.max_column + 1):
            for row_num in range(1, ws.max_row + 1):
                ws.cell(row=row_num, column=column_num).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Жирный шрифт верхней строчке + жирная полоска под
    for column_num in range(1, ws.max_column + 1):
        ws[get_column_letter(column_num) + '1'].font = Font(bold = True)
        ws[get_column_letter(column_num) + '1'].border = Border(top=thin, left=thin, right=thin, bottom=thick)

    # Сортировка
    def bubble_sort():
        # Устанавливаем swapped в True, чтобы цикл запустился хотя бы один раз
        swapped = True
        while swapped:
            swapped = False
            for i in range(2, ws.max_row):
                if ws['D' + str(i)].value > ws['D' + str(i+1)].value:
                    # Меняем элементы
                    for j in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                        ws[j + str(i)].value, ws[j + str(i+1)].value = ws[j + str(i+1)].value, ws[j + str(i)].value
                    # Устанавливаем swapped в True для следующей итерации
                    swapped = True

    bubble_sort()

    ws.auto_filter.ref = "D1:D100"
    ws.auto_filter.add_sort_condition("D2:D100")

    # Серый цвет через строчку
    greyFill = PatternFill(start_color='FFDDDDDD',
                        end_color='FFDDDDDD',
                        fill_type='solid')

    for row_num in range(1, ws.max_row + 1):
        if row_num % 2 == 0:
            for column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws[column_letter + str(row_num)].fill = greyFill

    # Если есть ПОЗДНИЙ ВЫЕЗД
    for row_num in range(2, ws.max_row + 1):
        if '12:00' not in ws['C' + str(row_num)].value:
            ws['I' + str(row_num)].value = 'ПВ'

    # ________Если есть файл за предыдущий день________
    if f'Living_{file_date_last}.xlsx' in os.listdir():
        print('[Проживание] Найден отчёт за вчерашний день, переношу нужную информацию.')
        file_name_last = f'Living_{file_date_last}.xlsx'
        file_path_last = os.getcwd() + '\\' + file_name_last
        print(f'[Проживание] Открываю вчерашний файл Living_{file_date_last}.xlsx.')
        wb_l = openpyxl.load_workbook(file_name_last)
        ws_l = wb_l.active
        # Проверить есть ли совпадения гостей и перенести информацию G, H, I
        for row_num_l in range(2, ws_l.max_row + 1):
            for row_num in range(2, ws.max_row + 1):        
                if ws_l['A' + str(row_num_l)].value == ws['A' + str(row_num)].value and ws_l['D' + str(row_num_l)].value == ws['D' + str(row_num)].value and ws_l['E' + str(row_num_l)].value == ws['E' + str(row_num)].value:
                    ws['G' + str(row_num)].value = ws_l['G' + str(row_num_l)].value
                    ws['H' + str(row_num)].value = ws_l['H' + str(row_num_l)].value
                    ws['I' + str(row_num)].value = ws_l['I' + str(row_num_l)].value
        # Ввести дополнительную информацию под табличкой
        n = 0
        while True:
            if ws_l['B' + str(ws_l.max_row - n)].value == None:
                ws['A' + str(ws.max_row + 1)].value = ws_l['A' + str(ws_l.max_row - n)].value
                n += 1
            else:
                break
        # Перенести старый отчёт в папку с предыдущими отчётами
        print('[Проживание] Переношу старый отчёт в папку old')
        try:
            os.rename(file_path_last, os.getcwd() + '\\old\\' + file_name_last)
        except:
            os.remove(file_path_last)

    # Сохранение файла
    print('[Проживание] Сохраняю файл.')
    wb.save(f'Living_{file_date}.xlsx')

    os.remove(file_path)

    print(f'[Проживание] ГОТОВО. Лист проживающих сохранён под именем "Living_{file_date}".')

def file_check():
    excel_files = []
    for i in range(len(os.listdir())):
        if os.listdir()[i][:4] == '4535':
            excel_files.append(os.listdir()[i])

    for i in range(len(excel_files)):
        if excel_files[i].endswith('xls'):
            kitchen(excel_files[i])
        elif excel_files[i].endswith('xlsx'):
            wb = openpyxl.load_workbook(excel_files[i])
            ws = wb.active
            dates_list = []
            for row_num in range(2, ws.max_row + 1):
                dates_list.append(ws['E' + str(row_num)].value)
            if len(set(dates_list)) == 1:
                checkin(excel_files[i])
            else:
                living(excel_files[i])

file_check()

input('\nГОТОВО. Для выхода нажмите любую клавишу.')