import os
import sys
import openpyxl
import datetime as dt
from openpyxl.styles import Font, Border, Side, NamedStyle, PatternFill, Color, colors, Alignment
from openpyxl.utils import get_column_letter
import win32com.client as win32
import zipfile
from tkinter import *
from tkinter import scrolledtext
from tkinter import messagebox 
from tkinter import filedialog
from tkinter import ttk  
from tkinter.ttk import Radiobutton 
from tkinter.ttk import Progressbar

class Report():
    """Класс для создания отчета"""

    def __init__(self, file_name):
        """Инициализировать отчет"""
        self.file_name = file_name
        self.file_date = ''
        self.name = ''

    def get_file_path(self):
        """Получает путь файла и открывает его для редактирования"""
        self.file_path = os.getcwd() + '\\' + self.file_name

    def delete_useless_columns(self, ws):
        """Удаляет лишние столбцы"""
        useless_columns = [1, 3, 4, 7, 8, 11, 13]

        for column_num in reversed(range(1, ws.max_column + 1)):
            if column_num in useless_columns:
                ws.delete_cols(column_num)
    
    def get_useful_columns(self, ws):
        """Получает список столбцов, с которыми нужно будет работать"""
        self.col_list = []
        for col_num in range(1, ws.max_column + 1):
            self.col_list.append(get_column_letter(col_num))

    def get_rooms_for_notes(self, ws):
        """Добавляет весь список комнат в словарь чтобы потом использовать номер комнаты как ключ и добавлять в его значения примечания"""
        self.notes = {}
        for row_num in range(2, ws.max_row + 1):
            room = ws['I' + str(row_num)].value
            self.notes[room] = ''

    def edit_columns_width(self, ws):
        """Редактирует ширину столбцов"""
        if len(self.col_list) == 8:
            col_width = [9.14, 5, 5, 7.86, 22.86, 4.29, 7, 27.86]
        elif len(self.col_list) == 9:
            col_width = [9.14, 5, 5, 7.14, 21.43, 4.29, 7, 6, 26.14]
        else:
            print(f'[{self.name}] ОШИБКА: Неправильное количество столбцов.')
            return

        for i in range(len(self.col_list)):
            ws.column_dimensions[self.col_list[i]].width = col_width[i]
    
    def add_border(self, ws):
        """Добавляет обрамление рамкой для таблицы"""
        thin = Side(border_style='thin', color='000000')

        for col_num in range(1, ws.max_column + 1):
            for row_num in range(1, ws.max_row + 1):
                ws.cell(row=row_num, column=col_num).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def edit_top_row(self, ws):
        """Добавляет жирный шрифт в верхней строчке + жирную полоску под"""
        thin = Side(border_style='thin', color='000000')
        thick = Side(border_style='thick', color='000000')

        for col_num in self.col_list:
            ws[col_num + '1'].font = Font(bold = True)
            ws[col_num + '1'].border = Border(top=thin, left=thin, right=thin, bottom=thick)
    
    def add_double_border(self, ws):
        """Добавляет двойную линию между номерами на 1, 2 и 3 этажах. Используется после сортировки."""
        double = Side(border_style='double', color = '000000')
        thin = Side(border_style='thin', color='000000')

        for row_num in range(2, ws.max_row):
            room = ws['D' + str(row_num)].value
            next_room = ws['D' + str(row_num + 1)].value

            if (next_room - room) > 50:
                for col_num in range(1, ws.max_column + 1):
                    ws[get_column_letter(col_num) + str(row_num)].border = Border(top=thin, left=thin, right=thin,bottom=double)

    def bubble_sort(self, ws):
        """Сортировка табличной части по номерам в порядке возрастания"""
        swapped = True
        while swapped:
            swapped = False
            for i in range(2, ws.max_row):
                if ws['D' + str(i)].value > ws['D' + str(i+1)].value:
                    # Меняем элементы
                    for j in self.col_list:
                        ws[j + str(i)].value, ws[j + str(i+1)].value = ws[j + str(i+1)].value, ws[j + str(i)].value
                    # Устанавливаем swapped в True для следующей итерации
                    swapped = True
    
    def add_sort(self, ws):
        """Добавляет сортировку по номеру"""
        ws.auto_filter.ref = "D2:D100"
        ws.auto_filter.add_sort_condition("D3:D100")

    def add_grey_rows(self, ws):
        """Добавляет серый цвет через строчку"""
        greyFill = PatternFill(start_color='FFDDDDDD',
                        end_color='FFDDDDDD',
                        fill_type='solid')

        for row_num in range(1, ws.max_row + 1):
            if row_num % 2 == 0:
                for column_letter in self.col_list:
                    ws[column_letter + str(row_num)].fill = greyFill

    def add_header(self, ws):
        """Добавляет заголовок с названием отчёта и датой"""
        blank = Side()

        ws.insert_rows(1)

        head_cell = ws['A1']

        head_cell.value = f'{self.name} {self.file_date}'
        head_cell.border = Border(top=blank, left=blank, right=blank, bottom=blank)
        head_cell.font = Font(sz = 24)
        head_cell.alignment = Alignment(horizontal="center", vertical="center")

        if len(self.col_list) == 8:
            ws.merge_cells('A1:H1') 
        elif len(self.col_list) == 9:
            ws.merge_cells('A1:I1') 
        else:
            print(f'[{self.name}] ОШИБКА: Неправильное количество столбцов.')
            return

    def delete_original_file(self):
        """Удаляет файл. Рекомендуется использовать только после сохранения"""
        os.remove(self.file_path)

class KitchenReport():
    """Класс для создания отчета кухни"""
    
    def __init__(self, file_name):
        """Инициализировать отчет кухни"""
        self.file_name = file_name
        self.fname = file_name
        self.name = 'Кухня'
        self.file_date = ''

    def get_file_path(self):
        """Получить имя и путь файла (file_name и file_path)"""
        self.file_name = self.file_name + 'x'
        self.file_path = os.getcwd() + '\\' + self.fname

    def convert_xls_xlsx(self):
        """Конвертирование расширение .xls -> .xlsx"""
        fname = self.file_path
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fname)
        wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
        wb.Close()                               #FileFormat = 56 is for .xls extension
        excel.Application.Quit()

    def get_date(self, ws):
        """Назначение даты файла"""
        self.file_date_y = (dt.date(int(ws['B2'].value[6:10]), int(ws['B2'].value[3:5]), int(ws['B2'].value[:2])) - dt.timedelta(1)).strftime('%d.%m.%Y')
        self.file_date = ws['B2'].value

    def editing_1(self, ws):
        """Редактирование"""
        ws.unmerge_cells('A4:B4')
        ws.delete_rows(1)
        ws['B3'] = ws['A3'].value   # Перенос 'Номер'
        ws['B' + str(ws.max_row)] = ''
        for row_num in range(3, ws.max_row):      # Очистка столбца 'A#'
            ws['A' + str(row_num)] = ''

    def count_guests(self, ws):
        """Подсчёт гостей"""
        for row_num in range(4, ws.max_row+1):
            try:      
                ws['C' + str(row_num)] = int(ws['D' + str(row_num)].value) + int(ws['E' + str(row_num)].value)
            except:
                ws['C' + str(row_num)] = ''
            try:
                ws['F' + str(row_num)] = int(ws['G' + str(row_num)].value) + int(ws['H' + str(row_num)].value)
            except:
                ws['F' + str(row_num)] = ''
            try:
                ws['I' + str(row_num)] = int(ws['J' + str(row_num)].value) + int(ws['K' + str(row_num)].value)
            except:
                ws['I' + str(row_num)] = ''

    def delete_useless_columns(self, ws):
        """Удаление лишних ячеек"""
        useless_columns = [4, 5, 7, 8, 10, 11]
        for column_num in reversed(range(1, ws.max_column + 1)):
            if column_num in useless_columns:
                ws.delete_cols(column_num)

        for row_num in range(4, ws.max_row):
            if ws['C' + str(row_num)].value == '' and ws['D' + str(row_num)].value == '' and ws['E' + str(row_num)].value == '':
                ws.delete_rows(row_num)

    def count_today_tomorrow_1(self, ws):
        """Подсчёт СЕГОДНЯ + ЗАВТРА 1"""
        self.today = ws['E' + str(ws.max_row)].value + ws['C' + str(ws.max_row)].value
        self.tomorrow = ws['E' + str(ws.max_row)].value + ws['D' + str(ws.max_row)].value
        self.guest_today = 'ГОСТ'
        self.guest_tomorrow = 'ГОСТ'

        if self.today in range(5, 21):
            self.guest_today += 'ЕЙ'
        elif self.today % 10 == 1:
            self.guest_today += 'Ь'
        elif 2 <= self.today % 10 <= 4:
            self.guest_today += 'Я'
        else:
            self.guest_today += 'ЕЙ'

        if self.tomorrow in range(5, 21):
            self.guest_tomorrow += 'ЕЙ'
        elif self.tomorrow % 10 == 1:
            self.guest_tomorrow += 'Ь'
        elif 2 <= self.tomorrow % 10 <= 4:
            self.guest_tomorrow += 'Я'
        else:
            self.guest_tomorrow += 'ЕЙ'

    def editing_2(self, ws):
        """Редактирование"""
        ws['E3'] = 'Прожив'
        ws['A' + str(ws.max_row)] = 'Номеров:' + str(ws.max_row - 4)
        ws['C' + str(ws.max_row)] = 'Выезд:' + str(ws['C' + str(ws.max_row)].value)
        ws['D' + str(ws.max_row)] = 'Заезд:' + str(ws['D' + str(ws.max_row)].value)
        ws['E' + str(ws.max_row)] = 'Прожив:' + str(ws['E' + str(ws.max_row)].value)

    def count_today_tomorrow_2(self, ws):
        """Подсчёт СЕГОДНЯ + ЗАВТРА 2"""
        ws['A' + str(ws.max_row+2)] = f'СЕГОДНЯ: {self.today} {self.guest_today}'
        ws['A' + str(ws.max_row+1)] = f'ЗАВТРА: ~{self.tomorrow} {self.guest_tomorrow}'

    def edit_columns_width(self, ws):
        """Редактирует ширину столбцов"""
        column_letter = ['B', 'C', 'D', 'E']
        column_width = [7.14, 9.28, 8.71, 10.71]
        for i in range(len(column_letter)):
            ws.column_dimensions[column_letter[i]].width = column_width[i]

    def edit_rows_width(self, ws):
        """Редактирует ширину строк"""
        for i in range(3, ws.max_row-3):
            ws.row_dimensions[i].height = 11.25

    def add_border(self, ws):
        """Добавляет обрамление рамкой"""
        dotted = Side(border_style='dotted', color='000000')
        for column_num in range(2, ws.max_column+1):
                for row_num in range(3, ws.max_row-3):
                    ws.cell(row=row_num, column=column_num).border = Border(top=dotted, left=dotted, right=dotted, bottom=dotted)
                    ws.cell(row=row_num, column=column_num).alignment = Alignment(horizontal="center", vertical="center")

    def edit_font_size(self, ws):
        """Изменяет размер шрифта"""
        ws['A1'].font = Font(size = 20)
        ws['B1'].font = Font(size = 20)
        ws['A' + str(ws.max_row)].font = Font(size = 20)
        ws['A' + str(ws.max_row-1)].font = Font(size = 20)

    def check_yesterday_file(self, ws):
        """Проверяет если есть файл за предыдущий день"""
        if f'Kitchen_{self.file_date_y}.xlsx' in os.listdir():
            file_name_y = f'Kitchen_{self.file_date_y}.xlsx'
            file_path_y = os.getcwd() + '\\' + file_name_y
            try:
                os.rename(file_path_y, os.getcwd() + '\\old\\' + file_name_y)
            except:
                os.remove(file_path_y)

    def save_file(self, wb):
        """Сохраняет файл"""
        wb.save(f'Kitchen_{self.file_date}.xlsx')

    def delete_original_file(self):
        """Удаляет файлы-источники"""
        os.remove(self.file_path)
        os.remove(self.file_path + 'x') 

class CheckInReport(Report):
    """Класс для создания отчета заезда"""

    def __init__(self, file_name):
        """Инициализировать отчет заезда"""
        super().__init__(file_name)
        self.name = 'Заезд'

    def get_date(self, ws):
        """Получает дату файла"""
        cell = ws['E2'].value
        self.file_date = cell[:10]
        self.file_date_y = (dt.date(int(cell[6:10]), int(cell[3:5]), int(cell[:2])) - dt.timedelta(1)).strftime('%d.%m.%Y')

    def add_rows(self, ws):
        """Добавляет 5 дополнительных ячеек для новых броней"""
        thin = Side(border_style='thin', color='000000')
        add_row = ws.max_row
        for column_num in range(1, ws.max_column + 1):
                for row_num in range(add_row + 1, add_row + 6):
                    ws.cell(row=row_num, column=column_num).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for i in range(ws.max_row-4, ws.max_row+1):
            ws.row_dimensions[i].height = 18.75
    
    def check_early_checkin(self, ws):
        """Добавляет в примечании пометку РЗ если есть ранний заезд. Используется перед удалением ненужных столбцов."""
        for row_num in range(2, ws.max_row):
            checkin = ws['E' + str(row_num)].value
            room = ws['I' + str(row_num)].value
            time = ws['E' + str(row_num)].value[-5:-3]

            if '14:00' not in checkin:
                self.notes[room] += 'РЗ~' + time + 'ч, '
    
    def check_breakfast(self, ws):
        """Добавляет в примечании пометку (б/з) если бронь Booking без завтрака. Используется перед удалением ненужных столбцов"""
        for row_num in range (2, ws.max_row):
            source = ws['B' + str(row_num)].value
            tariff = ws['G' + str(row_num)].value
            room = ws['I' + str(row_num)].value
            
            if source == "Booking.com" and 'без завтрака' in tariff:
                self.notes[room] += '(б/з), '

    def add_notes(self, ws):
        for row_num in range(2, ws.max_row + 1):
            room = ws['D' + str(row_num)].value
            ws['H' + str(row_num)].value = self.notes[room]

    def check_yesterday_file(self, ws):
        """Ищет в папке файл за предыдущий день. При его наличии переносит его в папку Old."""
        if f'Check-in_{self.file_date_y}.xlsx' in os.listdir():
            print('[Заезд] Найден отчёт за вчерашний день. Переношу старый отчёт в папку old.')
            file_name_y = f'Check-in_{self.file_date_y}.xlsx'
            file_path_y = os.getcwd() + '\\' + file_name_y
            try:
                os.rename(file_path_y, os.getcwd() + '\\old\\' + file_name_y)
            except:
                os.remove(file_path_y)

    def save_file(self, wb):
        """Сохраняет файл"""
        wb.save(f'Check-in_{self.file_date}.xlsx')

class LivingReport(Report):
    """Класс для создания отчета проживания"""

    def __init__(self, file_name):
        """Инициализировать отчет проживания"""
        super().__init__(file_name)
        self.name = 'Проживание'

    def get_date(self, ws, day, day_y):
        """Получает дату файла"""
        self.file_date = day
        self.file_date_y = day_y
        # day_today = dt.date.today().strftime('%d.%m.%Y')
        # day_yesterday = (dt.date.today() - dt.timedelta(1)).strftime('%d.%m.%Y')
        # day_tomorrow = (dt.date.today() + dt.timedelta(1)).strftime('%d.%m.%Y')

        # print(f'\nНа какое число делаем проживание? \n\n\
        #     Введите [1] если на сегодня {day_today}. \n\
        #     Введите [2] если на завтра {day_tomorrow}.')

        # n = int(input('\nВведите цифру и нажмите ENTER: '))

        # if n == 1:
        #     self.file_date = day_today
        #     self.file_date_y = day_yesterday
        # elif n == 2:
        #     self.file_date = day_tomorrow
        #     self.file_date_y = day_today
        # elif n == 3:
        #     another = input('\nВведите дату в формате ДД.ММ.ГГГГ: ')
        #     int_year = int(another.split('.')[2])
        #     int_month = int(another.split('.')[1])
        #     int_day = int(another.split('.')[0])
        #     day_another = dt.datetime(int_year, int_month, int_day)
        #     self.file_date = day_another.strftime('%d.%m.%Y')
        #     self.file_date_y = (day_another - dt.timedelta(1)).strftime('%d.%m.%Y')

    def get_rooms_for_notes(self, ws):
        """Добавляет весь список комнат в словарь чтобы потом использовать номер комнаты как ключ и добавлять в его значения примечания"""
        self.notes = {}
        for row_num in range(2, ws.max_row + 1):
            room = ws['I' + str(row_num)].value
            self.notes[room] = {'total' : '', 'minibar' : '', 'notes' : ''}

    def check_yesterday_file(self, ws):
        """Проверяет есть ли файл за предыдущий день. Если есть - добавляет в примечания информацию для тех номеров, которые есть во вчерашнем и новом списках + информацию под таблицей"""
        self.bottom_rows = []

        if f'Living_{self.file_date_y}.xlsx' in os.listdir():
            file_name_y = f'Living_{self.file_date_y}.xlsx'
            file_path_y = os.getcwd() + '\\' + file_name_y
            wb_y = openpyxl.load_workbook(file_name_y)
            ws_y = wb_y.active

            # Проверить есть ли совпадения гостей и перенести информацию Итого, мб, Примечания
            for row_num_y in range(3, ws_y.max_row + 1):
                source_y = ws_y['A' + str(row_num_y)].value
                room_y = ws_y['D' + str(row_num_y)].value
                guest_y = ws_y['E' + str(row_num_y)].value
                total_y = ws_y['G' + str(row_num_y)].value
                minibar_y = ws_y['H' + str(row_num_y)].value
                notes_y =  ws_y['I' + str(row_num_y)].value
                for row_num in range(2, ws.max_row + 1):
                    source = ws['B' + str(row_num)].value
                    room = ws['I' + str(row_num)].value
                    guest = ws['J' + str(row_num)].value
                    if source == source_y and room == room_y and guest == guest_y:
                        if total_y != None:
                            self.notes[room]['total'] = total_y
                        if minibar_y != None:
                            self.notes[room]['minibar'] = minibar_y
                        if notes_y  != None:
                            self.notes[room]['notes'] = notes_y 

            # Ввести дополнительную информацию под табличкой
            n = 0
            while True:
                if ws_y['B' + str(ws_y.max_row - n)].value == None:
                    self.bottom_rows.insert(0, ws_y['A' + str(ws_y.max_row - n)].value)
                    n += 1
                else:
                    break


            # Перенести старый отчёт в папку с предыдущими отчётами
            try:
                os.rename(file_path_y, os.getcwd() + '\\old\\' + file_name_y)
            except:
                os.remove(file_path_y)

    def check_late_checkout(self, ws):
        """Проверяет у гостей с пустыми примечаниями и добавляет к ним ПВ или (б/з) при необходимости"""
        for row_num in range (2, ws.max_row):
            source = ws['B' + str(row_num)].value
            tariff = ws['G' + str(row_num)].value
            room = ws['I' + str(row_num)].value
            checkout = ws['F' + str(row_num)].value
            
            if self.notes[room]['notes'] == '' and source == "Booking.com" and 'без завтрака' in tariff:
                self.notes[room]['notes'] += '(б/з), '

            if self.notes[room]['notes'] == '' and '12:00' not in checkout:
                self.notes[room]['notes'] += 'ПВ, '


    def add_notes(self, ws):
        for row_num in range(2, ws.max_row + 1):
            room = ws['D' + str(row_num)].value

            ws['I' + str(row_num)].value = self.notes[room]['notes']

            if self.notes[room]['total'] != '':
                ws['G' + str(row_num)].value = self.notes[room]['total']
            
            if self.notes[room]['minibar'] != '':
                ws['H' + str(row_num)].value = self.notes[room]['minibar']


    def add_bottom_notes(self, ws):
        """Выводит дополнительную информацию под таблицей"""
        while True:
            try:
                self.bottom_rows.remove(None)
            except:
                break
        for i in self.bottom_rows:
            if 'Смена белья:' in i:
                self.bottom_rows.remove(i)

        bottom_space = ws.max_row + 3
        for i in range(len(self.bottom_rows)):
            ws['A' + str(bottom_space + i)].value = self.bottom_rows[i]

    def check_linen_change(self, ws):
        """Добавляет список комнат, в которых нужно сменить бельё."""
        self.linen_rooms = []
        d_today = dt.datetime.strptime(self.file_date, '%d.%m.%Y')

        for row_num in range(2, ws.max_row + 1):
            checkin = ws['E' + str(row_num)].value[:10]
            checkout = ws['F' + str(row_num)].value[:10]
            room = ws['I' + str(row_num)].value
            d_in = dt.datetime.strptime(checkin, '%d.%m.%Y')
            d_out = dt.datetime.strptime(checkout, '%d.%m.%Y')
            d_amount = (d_out - d_in)

            clean_day = []
            if d_amount.days > 3:
                for i in range(2, d_amount.days, 2):
                    if (d_in + dt.timedelta(i+1)) < d_out:
                        clean_day.append(d_in + dt.timedelta(i))

                if d_today in clean_day:
                    self.linen_rooms.append(room)

        self.linen_rooms.sort()

    def add_linen_change(self, ws):
        """Выводит на экран список смены белья"""
        if len(self.linen_rooms) == 0:
            msg = 'Смена белья: -'
        else:
            msg = 'Смена белья: ' + ', '.join(map(str, self.linen_rooms))
        ws['A' + str(ws.max_row + 3)].value = msg

    def save_file(self, wb):
        """ Сохраняет файл """
        wb.save(f'Living_{self.file_date}.xlsx')

def make_kitchen_report(file_name):
    kitchen = KitchenReport(file_name)
    kitchen.get_file_path()
    kitchen.convert_xls_xlsx()

    wb = openpyxl.load_workbook(kitchen.file_name)
    ws = wb.active
    ws.title = 'Кухня'

    kitchen.get_date(ws)
    kitchen.editing_1(ws)
    kitchen.count_guests(ws)
    kitchen.delete_useless_columns(ws)

    # Пересохранение файла потому что вылезает ошибка, что строк больше, чем нужно
    # (баг возникает при наличии позднего/раннего выезда/заезда)
    kitchen.file_name = f'Kitchen_{kitchen.file_date}.xlsx'
    wb.save(kitchen.file_name)  
    wb = openpyxl.load_workbook(kitchen.file_name)
    ws = wb.active

    kitchen.count_today_tomorrow_1(ws)
    kitchen.editing_2(ws)
    kitchen.count_today_tomorrow_2(ws)
    kitchen.edit_columns_width(ws)
    kitchen.edit_rows_width(ws)
    kitchen.add_border(ws)
    kitchen.edit_font_size(ws)
    kitchen.check_yesterday_file(ws)
    kitchen.save_file(wb)
    kitchen.delete_original_file()

    print(f'[Кухня] ГОТОВО. Отчёт сохранён под именем "Kitchen_{kitchen.file_date}"')

def make_checkin_report(file_name):
    checkin = CheckInReport(file_name)
    checkin.get_file_path()

    wb = openpyxl.load_workbook(checkin.file_name)
    ws = wb.active
    ws.title = 'Заезд'

    checkin.get_date(ws)
    checkin.get_rooms_for_notes(ws)
    checkin.check_early_checkin(ws)
    checkin.check_breakfast(ws)
    checkin.delete_useless_columns(ws)

    ws['H1'] = 'Примечания'

    checkin.get_useful_columns(ws)
    checkin.edit_columns_width(ws)
    checkin.add_border(ws)
    checkin.edit_top_row(ws)
    checkin.bubble_sort(ws)
    checkin.add_sort(ws)
    checkin.add_double_border(ws)
    checkin.add_notes(ws)
    checkin.add_grey_rows(ws)
    checkin.check_yesterday_file(ws)
    checkin.add_header(ws)
    checkin.add_rows(ws)
    checkin.save_file(wb)
    checkin.delete_original_file()

    print(f'[Заезд] ГОТОВО. Отчёт сохранён под именем "Check-in_{checkin.file_date}"')

def make_living_report(file_name, day, day_y):
    living = LivingReport(file_name)
    living.get_file_path()

    wb = openpyxl.load_workbook(living.file_name)
    ws = wb.active
    ws.title = 'Проживание'
    
    living.get_date(ws, day, day_y)
    living.get_rooms_for_notes(ws)
    living.check_yesterday_file(ws) 
    living.check_late_checkout(ws) 
    living.check_linen_change(ws)
    living.delete_useless_columns(ws)

    ws['H1'] = 'мб'
    ws['I1'] = 'Примечания'

    living.get_useful_columns(ws)
    living.edit_columns_width(ws)
    living.add_border(ws)
    living.edit_top_row(ws)
    living.bubble_sort(ws)
    living.add_sort(ws)
    living.add_double_border(ws)
    living.add_grey_rows(ws)
    living.add_notes(ws)
    living.add_bottom_notes(ws)
    living.add_linen_change(ws)
    living.add_header(ws)
    living.save_file(wb)
    living.delete_original_file()

    print(f'[Проживание] ГОТОВО. Отчёт сохранён под именем "Living_{living.file_date}"')

def zip_file_check(download_folder):
    zip_files = []
    # download_folder = 'D:\\Downloads'
    # upload_folder = 'D:\\OneDrive\\Bnovo'
    upload_folder = os.getcwd()
    file_list = os.listdir(path = download_folder)
    for i in range(len(file_list)):
        if file_list[i].startswith('4535') and file_list[i].endswith('zip'):
            zip_files.append(file_list[i])

    if not zip_files:
        return False

    for i in range(len(zip_files)):
        file_zip = zipfile.ZipFile(download_folder + '\\' + zip_files[i])
        file_zip.extractall(upload_folder)
        file_zip.close()
        os.remove(download_folder + '\\' + zip_files[i])
    
    return True


def file_check(day, day_y):
    excel_files = []
    for i in range(len(os.listdir())):
        if os.listdir()[i][:4] == '4535':
            excel_files.append(os.listdir()[i])

    if not excel_files:
        return False

    for i in range(len(excel_files)):
        if excel_files[i].endswith('xls'):
            make_kitchen_report(excel_files[i])
        elif excel_files[i].endswith('xlsx'):
            wb = openpyxl.load_workbook(excel_files[i])
            ws = wb.active
            dates_list = []
            for row_num in range(2, ws.max_row + 1):
                dates_list.append(ws['E' + str(row_num)].value[:10])
            if len(set(dates_list)) == 1:
                make_checkin_report(excel_files[i])
            else:
                make_living_report(excel_files[i], day, day_y)

    return True

def main():
    def clicked():
        infotxt.configure(state ='normal')
        infotxt.insert(1.0, 'Запуск программы...\n')
        bar.start(10)

        # Обработка чекбокса
        if chk_state.get():
            infotxt.insert(1.0, 'Проверка папки на наличие архивов.\n')
            try:
                if zip_file_check(path.get()):
                    infotxt.insert(1.0, 'Разархивация прошла успешно.\n')
                else:
                    messagebox.showerror('Ошибка', 'В папке не найдено ни одного архива.')
                    return
            except:
                messagebox.showerror('Ошибка', 'Не найден путь к папке.')

        # Обработка радио-кнопки
        if selected.get() == 1:
            today = dt.date.today().strftime('%d.%m.%Y')
            yesterday = (dt.date.today() - dt.timedelta(1)).strftime('%d.%m.%Y')
        elif selected.get() == 2:
            today = (dt.date.today() + dt.timedelta(1)).strftime('%d.%m.%Y')
            yesterday = dt.date.today().strftime('%d.%m.%Y')
        else:
            messagebox.showerror('Ошибка', 'Не указана дата для отчета Проживание')
            bar.stop()
            return

        infotxt.insert(1.0, 'Начинаю редактировать отчёты.\n')

        if file_check(today, yesterday):
            infotxt.insert(1.0, 'Готово. Все отчеты отредактированы.\n')
        else:
            messagebox.showerror('Ошибка', 'В папке не найдено ни одного отчета.')
            return


        bar.stop()
        bar['value'] = 100
        infotxt.configure(state ='disabled')

    def get_path():
        path.configure(state ='normal')
        path.delete(0, END)
        directory = filedialog.askdirectory()
        path.insert(INSERT, directory)
        path.configure(state ='disabled')

    def check_chk():
        if chk_state.get():
            pathbtn.configure(state ='normal')
        else:
            pathbtn.configure(state ='disabled')

    # Окно
    window = Tk()
    window.title("Bnovo Report")
    # window.geometry('500x300')

    # Чекбокс разархивации zip
    chk_state = BooleanVar()  
    chk_state.set(True)  # задайте проверку состояния чекбокса 
    chk = Checkbutton(window, 
                    text='Разархивировать файлы из папки', 
                    font=("Century Gothic", 10), 
                    var=chk_state, command=check_chk)


    chk.grid(column=0, row=0, columnspan=5, sticky='w')

    # Указание каталога
    txt = Label(window, text="Путь:")  
    txt.grid(column=0, row=1, sticky='w')  

    path = Entry(window, 
                width=50)
    path.insert(INSERT, 'D:/Downloads')
    path.grid(column=1, row=1, columnspan=3, sticky='w')
    path.configure(state ='disabled')

    pathbtn = Button(window, text="Выбрать...", command=get_path)
    pathbtn.grid(column=5, row=1)


    # Заголовок для радио-кнопок
    lbl = Label(window, 
                text="Дата отчета Проживание (при наличии)", 
                font=("Century Gothic", 10))  

    lbl.grid(column=0, row=3, columnspan=5) 

    # Радио-кнопки
    selected = IntVar()  
    rad1 = Radiobutton(window, 
                    text=f'Сегодня {dt.date.today().strftime("%d.%m.%Y")}', 
                    value=1, 
                    variable=selected)  

    rad2 = Radiobutton(window, 
                    text=f'Завтра {(dt.date.today() + dt.timedelta(1)).strftime("%d.%m.%Y")}', 
                    value=2, 
                    variable=selected)

    rad1.grid(column=0, row=4, columnspan=2)  
    rad2.grid(column=2, row=4, columnspan=2) 

    # Текстовая область
    infotxt = scrolledtext.ScrolledText(window, 
                                    width=40, 
                                    height=6)

    infotxt.grid(column=0, row=6, columnspan=3) 
    infotxt.configure(state ='disabled')

    # Прогрессбар
    bar = Progressbar(window, length=320)
    bar.grid(column=0, row=7, columnspan=3)

    # Кнопка запуска процесса формирования отчетов
    btn = Button(window,
                command=clicked, 
                text="Сделать \nотчеты", 
                bg='palegreen', 
                fg='black', 
                font=("Century Gothic", 12))

    btn.grid(column=5, row=6)

    window.mainloop()

main()